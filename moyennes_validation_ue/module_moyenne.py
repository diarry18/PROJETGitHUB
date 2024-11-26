import os
import openpyxl 


def calcul_moyennes_notes_UE1_S1():
    """
    Calcule la moyenne des notes de l'UE1 S1 pour chaque étudiant à partir des fichiers Excel contenant les notes.
    Retourne un dictionnaire contenant les moyennes de chaque étudiant pour l'UE1 S1, un autre dictionnaire contenant les moyennes de chaque étudiant pour l'UE1 S2, un dictionnaire contenant le résultat (validé ou non) de chaque étudiant pour l'UE1 et une liste de listes contenant les données à afficher dans le tableau HTML.
    """
    # Charger les fichiers Excel et initialisation des variables
    donnees_UE1 = {}
    resultat = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE1_S1 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE1_S1 = os.path.join(dossier_UE1_S1, "SAE_105_BILLON_BOUGUEREAU-main" , "data", "excel_notes" , "notes_S1" , "UE1.1") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE1_S1):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS1"
            fichier_excel = os.path.join(dossier_UE1_S1, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE1:
                    donnees_UE1[nom_eleve] += valeur
                else:
                    donnees_UE1[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE1.items():
        donnees_UE1[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Charger les fichiers Excel et initialisation des variables
    donnees_UE1_S2 = {}
    etatUE1 = {}
    resultat = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE1_S2 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE1_S2 = os.path.join(dossier_UE1_S2, "SAE_105_BILLON_BOUGUEREAU-main" , "data", "excel_notes" , "notes_S2" , "UE1") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE1_S2):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS2"
            fichier_excel = os.path.join(dossier_UE1_S2, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE1_S2:
                    donnees_UE1_S2[nom_eleve] += valeur
                else:
                    donnees_UE1_S2[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur           

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE1_S2.items():
        donnees_UE1_S2[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Calculer la moyenne finale et déterminer si l'UE est valide ou non
    for nom_eleve in donnees_UE1.keys():#.keys sert a cree une liste de nom dans un dictionnaire  ( ue3 est le dictionnaire et les nom sont les clef )
        if (donnees_UE1[nom_eleve] + donnees_UE1_S2[nom_eleve]) / 2 >= 10:
            resultat[nom_eleve] = "Valider"
        else:
            resultat[nom_eleve] = "Non valider"

    column_titles = ["Nom", "Prénom", "Moyenne UE1 S1", "Moyenne UE1 S2", "Résultat"]
    data_rows = []
    for nom_eleve in donnees_UE1.keys():
        nom, prenom = nom_eleve.split(' ')
        data_rows.append([nom, prenom, donnees_UE1[nom_eleve], donnees_UE1_S2[nom_eleve], resultat[nom_eleve]])        


    return donnees_UE1, donnees_UE1_S2, resultat, data_rows



def calcul_moyennes_notes_UE2_S1():
    """
    Calcule la moyenne des notes de l'UE2 S1 pour chaque étudiant à partir des fichiers Excel contenant les notes.
    Retourne un dictionnaire contenant les moyennes de chaque étudiant pour l'UE2 S1, un autre dictionnaire contenant les moyennes de chaque étudiant pour l'UE2 S2, un dictionnaire contenant le résultat (validé ou non) de chaque étudiant pour l'UE2 et une liste de listes contenant les données à afficher dans le tableau HTML.
    """
    # Charger les fichiers Excel et initialisation des variables
    donnees_UE2 = {}
    resultat = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE2_S1 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE2_S1 = os.path.join(dossier_UE2_S1, "SAE_105_BILLON_BOUGUEREAU-main" , "data", "excel_notes" , "notes_S1" , "UE1.2") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE2_S1):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS1"
            fichier_excel = os.path.join(dossier_UE2_S1, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE2:
                    donnees_UE2[nom_eleve] += valeur
                else:
                    donnees_UE2[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE2.items():
        donnees_UE2[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Charger les fichiers Excel et initialisation des variables
    donnees_UE2_S2 = {}
    etatUE2 = {}
    resultat_UE2 = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE2_S2 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE2_S2 = os.path.join(dossier_UE2_S2, "SAE_105_BILLON_BOUGUEREAU-main" , "data", "excel_notes" , "notes_S2" , "UE2") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE2_S2):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS2"
            fichier_excel = os.path.join(dossier_UE2_S2, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE2_S2:
                    donnees_UE2_S2[nom_eleve] += valeur
                else:
                    donnees_UE2_S2[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur           

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE2_S2.items():
        donnees_UE2_S2[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Calculer la moyenne finale et déterminer si l'UE est valide ou non
    for nom_eleve in donnees_UE2.keys():
        if (donnees_UE2[nom_eleve] + donnees_UE2_S2[nom_eleve]) / 2 >= 10:
            resultat_UE2[nom_eleve] = "Valider"
        else:
            resultat_UE2[nom_eleve] = "Non valider"

    column_titlesUE2 = ["Moyenne UE2 S1", "Moyenne UE2 S2", "Resultat UE2"]
    data_rowsUE2 = []
    for nom_eleve in donnees_UE2.keys():
        nom, prenom = nom_eleve.split(' ')
        data_rowsUE2.append([nom, prenom, donnees_UE2[nom_eleve], donnees_UE2_S2[nom_eleve], resultat_UE2[nom_eleve]])        


    return donnees_UE2, donnees_UE2_S2, resultat_UE2, data_rowsUE2



def calcul_moyennes_notes_UE3_S1():
    """
    Calcule la moyenne des notes de l'UE3 S1 pour chaque étudiant à partir des fichiers Excel contenant les notes.
    Retourne un dictionnaire contenant les moyennes de chaque étudiant pour l'UE3 S1, un autre dictionnaire contenant les moyennes de chaque étudiant pour l'UE3 S2, un dictionnaire contenant le résultat (validé ou non) de chaque étudiant pour l'UE3 et une liste de listes contenant les données à afficher dans le tableau HTML.
    """
    # Charger les fichiers Excel et initialisation des variables
    donnees_UE3 = {}
    resultat = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE3_S1 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE3_S1 = os.path.join(dossier_UE3_S1, "SAE_105_BILLON_BOUGUEREAU-main" , "data", "excel_notes" , "notes_S1" , "UE1.3") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE3_S1):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS1"
            fichier_excel = os.path.join(dossier_UE3_S1, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE3:
                    donnees_UE3[nom_eleve] += valeur
                else:
                    donnees_UE3[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE3.items():
        donnees_UE3[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Charger les fichiers Excel et initialisation des variables
    donnees_UE3_S2 = {}
    etatUE3 = {}
    resultat_UE3 = {}
    emplacementcode = os.getcwd() #determine l'emplacement du script dans l'arboressence 
    dossier_UE3_S2 = os.path.dirname(os.path.dirname(emplacementcode))
    dossier_UE3_S2 = os.path.join(dossier_UE3_S2, "SAE_105_BILLON_BOUGUEREAU-main" , "data", "excel_notes" , "notes_S2" , "UE3") #permet d'acceder au fichier sur n'importe quelle pc
    num_files = 0
    for nomfichier in os.listdir(dossier_UE3_S2):
        if nomfichier.endswith(".xlsx"): #chereche TOUT les fichier en .xlsx du dossier et les charge 
            num_files += 1 #"compte le nombre de fichier .xlsx dans dossierS2"
            fichier_excel = os.path.join(dossier_UE3_S2, nomfichier) #slelectione les "worksheet active"
            classeur = openpyxl.load_workbook(fichier_excel) # contient plusieur worksheet

            # Sélectionner la feuille de calcul
            feuille = classeur.active

            # Lire et traiter les donnes
            for ligne in feuille.iter_rows(min_row=2, max_row=feuille.max_row, values_only=True):
                # `ligne` = valeurs de chaque cellule dans la ligne
                nom_eleve = ligne[1] + ' ' + ligne[2]
                valeur = ligne[3]
                if nom_eleve in donnees_UE3_S2:
                    donnees_UE3_S2[nom_eleve] += valeur
                else:
                    donnees_UE3_S2[nom_eleve] = valeur
        
            classeur.close()  # Ferme le classeur           

    # divise la somme des notes par le nombre de fichier pour obtenir la moyenne
    for nom_eleve, somme in donnees_UE3_S2.items():
        donnees_UE3_S2[nom_eleve] = round(somme / num_files, 2) #calcule la moyenne et l'arrondie 

    # Calculer la moyenne finale et déterminer si l'UE est valide ou non
    for nom_eleve in donnees_UE3.keys():
        if (donnees_UE3[nom_eleve] + donnees_UE3_S2[nom_eleve]) / 2 >= 10:
            resultat_UE3[nom_eleve] = "Valider"
        else:
            resultat_UE3[nom_eleve] = "Non valider"

    column_titlesUE3 = ["Moyenne UE3 S1", "Moyenne UE3 S2", "Resultat UE3"]
    data_rowsUE3 = []
    for nom_eleve in donnees_UE3.keys():#.keys sert a cree une liste de nom dans un dictionnaire  ( ue3 est le dictionnaire et les nom sont les clef )
        nom, prenom = nom_eleve.split(' ')
        data_rowsUE3.append([nom, prenom, donnees_UE3[nom_eleve], donnees_UE3_S2[nom_eleve], resultat_UE3[nom_eleve]])        


    return donnees_UE3, donnees_UE3_S2, resultat_UE3, data_rowsUE3



def generate_html_file(file_name, title, column_titles,  column_titlesUE2 , column_titlesUE3 , data_rowsUE3 ,data_rows , data_rowsUE2):
    """
    Génère un fichier HTML contenant un tableau avec les données passées en paramètres.
    :parametre file_name: Le nom du fichier HTML à générer.
    :parametre title: Le titre de la page HTML.
    :parametre column_titles: Une liste contenant les titres des colonnes pour l'UE1.
    :paramerte column_titlesUE2: Une liste contenant les titres des colonnes pour l'UE2.
    :parametre column_titlesUE3: Une liste contenant les titres des colonnes pour l'UE3.
    :parametre data_rowsUE3: Une liste de listes contenant les données à afficher pour l'UE3.
    :parametre data_rows: Une liste de listes contenant les données à afficher pour l'UE1.
    :parametre data_rowsUE2: Une liste de listes contenant les données à afficher pour l'UE2.
    """
    #Contenu HTML
    file_name="moyennes_validation_UE.html"
    html_content = f"""
    <!DOCTYPE html>
    <html lang="fr">
    <head>
        <meta charset="utf8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>SAE 105-Traiter des donnees</title>
 <link href="../html/global.css" rel="stylesheet">
    </head>
    <body>
        <h1>{title}</h1>
        <table>
            <tr>
                {''.join(f'<th>{col}</th>' for col in column_titles)}
            </tr>
            {''.join(f'<tr>{"".join(f"<td>{data}</td>" for data in row)}</tr>' for row in data_rows )}
            <tr>
                {''.join(f'<th>{col}</th>' for col in column_titlesUE2)}
            </tr>
            {''.join(f'<tr>{"".join(f"<td>{data}</td>" for data in row)}</tr>' for row in data_rowsUE2 )}
            <tr>
                {''.join(f'<th>{col}</th>' for col in column_titlesUE3)}
            </tr>
            {''.join(f'<tr>{"".join(f"<td>{data}</td>" for data in row)}</tr>' for row in data_rowsUE3 )}

        </table>
    </body>
    </html>
    """

    #Ecriture du contenu dans le fichier spécifié
    with open(file_name, "w") as file:
        file.write(html_content)

    print(f"Le fichier {file_name} a été généré avec succès.")
